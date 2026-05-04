"""
Database models and operations for the Stock Market Scanner
"""
import json
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any
from pathlib import Path
import hashlib
import pandas as pd
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Text, Boolean
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.sql import func
from loguru import logger

from core.config import settings, DatabaseConfig

Base = declarative_base()


class User(Base):
    """Application user model"""
    __tablename__ = "users"

    id = Column(Integer, primary_key=True, index=True)
    email = Column(String(255), unique=True, index=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    full_name = Column(String(255), nullable=True)
    is_active = Column(Boolean, default=True)
    is_verified = Column(Boolean, default=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class RefreshSession(Base):
    """Refresh token sessions for rotation/revocation"""
    __tablename__ = "refresh_sessions"

    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, index=True, nullable=False)
    token_hash = Column(String(128), unique=True, index=True, nullable=False)
    user_agent = Column(String(512), nullable=True)
    ip_address = Column(String(64), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    expires_at = Column(DateTime, nullable=False)
    revoked_at = Column(DateTime, nullable=True)
    replaced_by = Column(String(128), nullable=True)


class Scan(Base):
    """Scan session model"""
    __tablename__ = "scans"
    
    id = Column(Integer, primary_key=True, index=True)
    timestamp = Column(DateTime, default=datetime.utcnow, index=True)
    status = Column(String(20), default="running")  # running, completed, failed
    total_stocks = Column(Integer, default=0)
    successful_stocks = Column(Integer, default=0)
    failed_stocks = Column(Integer, default=0)
    execution_time = Column(Float, default=0.0)
    strategies_run = Column(Text)  # JSON list of strategy names
    results_summary = Column(Text)  # JSON summary of results
    created_at = Column(DateTime, default=datetime.utcnow)
    completed_at = Column(DateTime, nullable=True)


class Strategy(Base):
    """Strategy configuration model"""
    __tablename__ = "strategies"
    
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(100), unique=True, index=True)
    display_name = Column(String(200))
    description = Column(Text)
    enabled = Column(Boolean, default=True)
    priority = Column(Integer, default=0)
    parameters = Column(Text)  # JSON parameters
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class ScanResult(Base):
    """Individual scan result model"""
    __tablename__ = "scan_results"
    
    id = Column(Integer, primary_key=True, index=True)
    scan_id = Column(Integer, index=True)
    strategy_name = Column(String(100), index=True)
    symbol = Column(String(20), index=True)
    current_price = Column(Float)
    result_data = Column(Text)  # JSON with strategy-specific data
    score = Column(Float, nullable=True)  # Calculated score for ranking
    created_at = Column(DateTime, default=datetime.utcnow)


class Stock(Base):
    """Stock information model"""
    __tablename__ = "stocks"
    
    id = Column(Integer, primary_key=True, index=True)
    symbol = Column(String(20), unique=True, index=True)
    name = Column(String(200))
    sector = Column(String(100), nullable=True)
    market_cap = Column(Float, nullable=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class StockData(Base):
    """Historical stock data model"""
    __tablename__ = "stock_data"
    
    id = Column(Integer, primary_key=True, index=True)
    symbol = Column(String(20), index=True)
    date = Column(DateTime, index=True)
    open_price = Column(Float)
    high_price = Column(Float)
    low_price = Column(Float)
    close_price = Column(Float)
    volume = Column(Integer)
    adjusted_close = Column(Float, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class DatabaseManager:
    """Database operations manager"""
    
    def __init__(self, database_url: str = None):
        self.database_url = database_url or settings.DATABASE_URL
        self.engine = create_engine(self.database_url, echo=False)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        
    def create_tables(self):
        """Create all database tables"""
        Base.metadata.create_all(bind=self.engine)
        self._create_indexes()
        
    def _create_indexes(self):
        """Create database indexes for performance"""
        with self.engine.connect() as conn:
            for index_sql in DatabaseConfig.INDEXES:
                try:
                    conn.execute(index_sql)
                except Exception as e:
                    logger.warning(f"Could not create index: {e}")
    
    def get_session(self) -> Session:
        """Get database session"""
        return self.SessionLocal()
    
    # Users
    def get_user_by_email(self, email: str) -> Optional[User]:
        session = self.get_session()
        try:
            return session.query(User).filter(User.email == email).first()
        finally:
            self.close_session(session)

    def create_user(self, email: str, password_hash: str, full_name: Optional[str] = None, is_verified: bool = False) -> int:
        session = self.get_session()
        try:
            user = User(email=email, password_hash=password_hash, full_name=full_name, is_verified=is_verified)
            session.add(user)
            session.commit()
            session.refresh(user)
            return user.id
        finally:
            self.close_session(session)

    # Refresh sessions
    def create_refresh_session(self, user_id: int, token_hash: str, expires_at: datetime, user_agent: Optional[str], ip_address: Optional[str]) -> int:
        session = self.get_session()
        try:
            rs = RefreshSession(user_id=user_id, token_hash=token_hash, expires_at=expires_at, user_agent=user_agent, ip_address=ip_address)
            session.add(rs)
            session.commit()
            session.refresh(rs)
            return rs.id
        finally:
            self.close_session(session)

    def get_refresh_session(self, token_hash: str) -> Optional[RefreshSession]:
        session = self.get_session()
        try:
            return session.query(RefreshSession).filter(RefreshSession.token_hash == token_hash).first()
        finally:
            self.close_session(session)

    def revoke_refresh_session(self, token_hash: str, replaced_by: Optional[str] = None):
        session = self.get_session()
        try:
            rs = session.query(RefreshSession).filter(RefreshSession.token_hash == token_hash).first()
            if rs and not rs.revoked_at:
                rs.revoked_at = datetime.utcnow()
                rs.replaced_by = replaced_by
                session.commit()
        finally:
            self.close_session(session)

    def revoke_all_sessions_for_user(self, user_id: int):
        session = self.get_session()
        try:
            session.query(RefreshSession).filter(RefreshSession.user_id == user_id, RefreshSession.revoked_at.is_(None)).update({RefreshSession.revoked_at: datetime.utcnow()})
            session.commit()
        finally:
            self.close_session(session)

    # Stocks CRUD
    def list_stocks(self, limit: int = 50, offset: int = 0) -> List[Dict[str, Any]]:
        session = self.get_session()
        try:
            q = session.query(Stock).order_by(Stock.symbol).limit(limit).offset(offset)
            items = q.all()
            return [
                {
                    'id': it.id,
                    'symbol': it.symbol,
                    'name': it.name,
                    'sector': it.sector,
                    'market_cap': it.market_cap,
                    'is_active': it.is_active,
                    'created_at': it.created_at,
                    'updated_at': it.updated_at,
                }
                for it in items
            ]
        finally:
            self.close_session(session)
    
    def create_stock(self, data: Dict[str, Any]) -> int:
        session = self.get_session()
        try:
            stock = Stock(**data)
            session.add(stock)
            session.commit()
            session.refresh(stock)
            return stock.id
        finally:
            self.close_session(session)
    
    def update_stock(self, stock_id: int, data: Dict[str, Any]) -> bool:
        session = self.get_session()
        try:
            stock = session.query(Stock).filter(Stock.id == stock_id).first()
            if not stock:
                return False
            for k, v in data.items():
                setattr(stock, k, v)
            session.commit()
            return True
        finally:
            self.close_session(session)
    
    def delete_stock(self, stock_id: int) -> bool:
        session = self.get_session()
        try:
            deleted = session.query(Stock).filter(Stock.id == stock_id).delete()
            session.commit()
            return deleted > 0
        finally:
            self.close_session(session)
    
    def get_active_symbols(self) -> List[str]:
        """Return active symbols as exchange-qualified strings (e.g., RELIANCE.NS)."""
        session = self.get_session()
        try:
            items = session.query(Stock).filter(Stock.is_active == True).all()
            symbols = []
            for it in items:
                sym = it.symbol.strip().upper() if it.symbol else ''
                if not sym:
                    continue
                # Append NSE suffix if not present
                symbols.append(sym if sym.endswith('.NS') else f"{sym}.NS")
            return symbols
        finally:
            self.close_session(session)
    
    def close_session(self, session: Session):
        """Close database session"""
        session.close()
    
    def save_scan(self, scan_data: Dict[str, Any]) -> int:
        """Save scan session and return scan ID"""
        session = self.get_session()
        try:
            scan = Scan(**scan_data)
            session.add(scan)
            session.commit()
            session.refresh(scan)
            return scan.id
        finally:
            self.close_session(session)
    
    def update_scan(self, scan_id: int, **kwargs):
        """Update scan session"""
        session = self.get_session()
        try:
            scan = session.query(Scan).filter(Scan.id == scan_id).first()
            if scan:
                for key, value in kwargs.items():
                    setattr(scan, key, value)
                session.commit()
        finally:
            self.close_session(session)
    
    def save_scan_results(self, scan_id: int, results: List[Dict[str, Any]]):
        """Save scan results"""
        session = self.get_session()
        try:
            for result in results:
                scan_result = ScanResult(
                    scan_id=scan_id,
                    strategy_name=result.get('strategy_name'),
                    symbol=result.get('symbol'),
                    current_price=result.get('current_price'),
                    result_data=json.dumps(result.get('result_data', {})),
                    score=result.get('score')
                )
                session.add(scan_result)
            session.commit()
        finally:
            self.close_session(session)
    
    def get_recent_scans(self, limit: int = 10) -> List[Dict[str, Any]]:
        """Get recent scan sessions"""
        session = self.get_session()
        try:
            scans = session.query(Scan).order_by(Scan.timestamp.desc()).limit(limit).all()
            return [
                {
                    'id': scan.id,
                    'timestamp': scan.timestamp,
                    'status': scan.status,
                    'total_stocks': scan.total_stocks,
                    'successful_stocks': scan.successful_stocks,
                    'execution_time': scan.execution_time,
                    'strategies_run': json.loads(scan.strategies_run) if scan.strategies_run else []
                }
                for scan in scans
            ]
        finally:
            self.close_session(session)
    
    def get_scan_results(self, scan_id: int, strategy_name: str = None) -> List[Dict[str, Any]]:
        """Get scan results for a specific scan"""
        session = self.get_session()
        try:
            query = session.query(ScanResult).filter(ScanResult.scan_id == scan_id)
            if strategy_name:
                query = query.filter(ScanResult.strategy_name == strategy_name)
            
            results = query.all()
            return [
                {
                    'id': result.id,
                    'strategy_name': result.strategy_name,
                    'symbol': result.symbol,
                    'current_price': result.current_price,
                    'result_data': json.loads(result.result_data) if result.result_data else {},
                    'score': result.score,
                    'created_at': result.created_at
                }
                for result in results
            ]
        finally:
            self.close_session(session)
    
    def get_strategy_performance(self, days: int = 30) -> Dict[str, Any]:
        """Get strategy performance statistics"""
        session = self.get_session()
        try:
            since_date = datetime.utcnow() - timedelta(days=days)
            
            # Get scan counts by strategy
            strategy_stats = session.query(
                ScanResult.strategy_name,
                func.count(ScanResult.id).label('total_results'),
                func.count(func.distinct(ScanResult.symbol)).label('unique_stocks')
            ).filter(
                ScanResult.created_at >= since_date
            ).group_by(ScanResult.strategy_name).all()
            
            return {
                'period_days': days,
                'strategies': [
                    {
                        'name': stat.strategy_name,
                        'total_results': stat.total_results,
                        'unique_stocks': stat.unique_stocks
                    }
                    for stat in strategy_stats
                ]
            }
        finally:
            self.close_session(session)

    def get_top_stocks(self, days: int = 30, strategy: Optional[str] = None, limit: int = 10) -> List[Dict[str, Any]]:
        """Get top stocks by frequency (and avg score) within the last N days, optionally filtered by strategy."""
        session = self.get_session()
        try:
            since_date = datetime.utcnow() - timedelta(days=days)
            query = session.query(
                ScanResult.symbol.label('symbol'),
                func.count(ScanResult.id).label('hits'),
                func.avg(ScanResult.score).label('avg_score'),
                func.max(ScanResult.created_at).label('last_seen')
            ).filter(ScanResult.created_at >= since_date)
            if strategy:
                query = query.filter(ScanResult.strategy_name == strategy)
            rows = query.group_by(ScanResult.symbol).order_by(func.count(ScanResult.id).desc(), func.max(ScanResult.created_at).desc()).limit(limit).all()
            return [
                {
                    'symbol': r.symbol,
                    'hits': int(r.hits or 0),
                    'avg_score': float(r.avg_score) if r.avg_score is not None else None,
                    'last_seen': r.last_seen
                }
                for r in rows
            ]
        finally:
            self.close_session(session)
    
    def cleanup_old_data(self, days: int = None):
        """Clean up old data to manage database size"""
        if days is None:
            days = settings.DATA_RETENTION_DAYS
        
        cutoff_date = datetime.utcnow() - timedelta(days=days)
        session = self.get_session()
        
        try:
            # Delete old scan results
            old_results = session.query(ScanResult).filter(
                ScanResult.created_at < cutoff_date
            ).delete()
            
            # Delete old scans
            old_scans = session.query(Scan).filter(
                Scan.timestamp < cutoff_date
            ).delete()
            
            # Delete old stock data
            old_stock_data = session.query(StockData).filter(
                StockData.date < cutoff_date
            ).delete()
            
            session.commit()
            
            return {
                'deleted_results': old_results,
                'deleted_scans': old_scans,
                'deleted_stock_data': old_stock_data
            }
        finally:
            self.close_session(session)
    
    def delete_scan(self, scan_id: int) -> Dict[str, int]:
        """Delete a scan and its associated results"""
        session = self.get_session()
        try:
            deleted_results = session.query(ScanResult).filter(ScanResult.scan_id == scan_id).delete()
            deleted_scans = session.query(Scan).filter(Scan.id == scan_id).delete()
            session.commit()
            return {"deleted_results": deleted_results, "deleted_scans": deleted_scans}
        finally:
            self.close_session(session)
    
    def export_scan_to_excel(self, scan_id: int, filename: str = None) -> str:
        """Export scan results to Excel file"""
        if filename is None:
            filename = f"scan_results_{scan_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = Path("./exports") / filename
        filepath.parent.mkdir(exist_ok=True)
        
        # Get scan info
        session = self.get_session()
        try:
            scan = session.query(Scan).filter(Scan.id == scan_id).first()
            if not scan:
                raise ValueError(f"Scan {scan_id} not found")
            
            # Get results by strategy
            results_by_strategy = {}
            strategies = json.loads(scan.strategies_run) if scan.strategies_run else []
            
            for strategy in strategies:
                results = self.get_scan_results(scan_id, strategy)
                if results:
                    results_by_strategy[strategy] = results
            
            # Create Excel file
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Summary sheet
                summary_data = []
                for strategy, results in results_by_strategy.items():
                    summary_data.append({
                        'Strategy': strategy.replace('_', ' ').title(),
                        'Stocks Found': len(results),
                        'Total Stocks Scanned': scan.total_stocks,
                        'Success Rate %': round((len(results) / scan.total_stocks) * 100, 2) if scan.total_stocks > 0 else 0
                    })
                
                summary_df = pd.DataFrame(summary_data)
                if not summary_df.empty:
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                else:
                    pd.DataFrame([{"Note": "No results available for this scan."}]).to_excel(writer, sheet_name='Summary', index=False)
                
                # Style Summary
                ws = writer.sheets['Summary']
                # Header styling
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1890FF")
                    cell.alignment = Alignment(horizontal="center")
                ws.freeze_panes = "A2"
                # Auto column widths
                for col_idx in range(1, ws.max_column + 1):
                    max_len = 0
                    col_letter = get_column_letter(col_idx)
                    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                        val = row[0]
                        max_len = max(max_len, len(str(val)) if val is not None else 0)
                    ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 50)

                # Individual strategy sheets
                for strategy, results in results_by_strategy.items():
                    if results:
                        df = pd.DataFrame(results)
                        # Convert result_data from JSON to columns
                        if 'result_data' in df.columns:
                            result_data_df = pd.json_normalize(df['result_data'])
                            df = pd.concat([df.drop('result_data', axis=1), result_data_df], axis=1)
                        
                        # Order and beautify columns
                        col_map = {
                            'strategy_name': 'Strategy',
                            'symbol': 'Symbol',
                            'current_price': 'Current Price',
                            'score': 'Score',
                        }
                        df.rename(columns=col_map, inplace=True)
                        leading_cols = [c for c in ['Strategy','Symbol','Current Price','Score'] if c in df.columns]
                        other_cols = [c for c in df.columns if c not in leading_cols]
                        df = df[leading_cols + other_cols]

                        sheet_name = strategy.replace('_', ' ').title()[:31]
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                        ws = writer.sheets[sheet_name]
                        # Header styling
                        for cell in ws[1]:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill("solid", fgColor="52C41A" if strategy != 'avwap_proximity' else "1890FF")
                            cell.alignment = Alignment(horizontal="center")
                        ws.freeze_panes = "A2"

                        # Auto column widths and basic number formats
                        headers = [cell.value for cell in ws[1]]
                        for col_idx, header in enumerate(headers, start=1):
                            col_letter = get_column_letter(col_idx)
                            max_len = len(str(header)) if header else 0
                            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True):
                                val = row[0]
                                max_len = max(max_len, len(str(val)) if val is not None else 0)
                            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 50)

                            # Set number formats
                            header_lower = (header or '').lower()
                            if 'price' in header_lower or header_lower in ['avwap', '52w high', '52w low']:
                                for cell in ws[col_letter][1:]:
                                    cell.number_format = '#,##0.00'
                            if '%' in header_lower or 'percent' in header_lower or 'return' in header_lower or 'difference' in header_lower or 'success rate' in header_lower:
                                for cell in ws[col_letter][1:]:
                                    cell.number_format = '0.00%'
                            if 'volume' in header_lower:
                                for cell in ws[col_letter][1:]:
                                    cell.number_format = '#,##0'
            
            return str(filepath)
        finally:
            self.close_session(session)


# Global database manager instance
db_manager = DatabaseManager()


def init_db():
    """Initialize database tables"""
    db_manager.create_tables()
    # Seed stocks from StrategyConfig if table is empty
    try:
        from core.config import StrategyConfig
        session = db_manager.get_session()
        try:
            count = session.query(Stock).count()
            if count == 0:
                for sym in StrategyConfig.NSE_SYMBOLS:
                    s = Stock(symbol=sym, name=None, sector=None, market_cap=None, is_active=True)
                    session.add(s)
                session.commit()
        finally:
            db_manager.close_session(session)
    except Exception as e:
        logger.warning(f"Could not seed stocks: {e}")
